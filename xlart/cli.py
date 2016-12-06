#!/usr/bin/env python2.7
import sys
from argparse import ArgumentParser

from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill

COLUMN_WIDTH = 2
ROW_HEIGHT = 15


def clamp(x):
    return max(0, min(x, 255))


def main(args=None):
    if args is None:
        args = sys.argv[1:]

    parser = ArgumentParser(description='Excel art generator')
    parser.add_argument('source', metavar='FILE.jpg|FILE.png', type=str,
                        help='The source image')
    parser.add_argument('destination', metavar='FILE.xlsx', type=str,
                        help='The excel file to create. (WILL BE OVERWRITTEN)')
    parser.add_argument('--scale', dest='scale', action='store',
                        nargs='?', type=int, default=2,
                        help='pixel downsampling (default: 2)')
    args = parser.parse_args()

    wb = Workbook()
    worksheet = wb.active

    img = Image.open(args.source)
    pix = img.load()
    (width, height) = img.size

    dx = 1
    for x in xrange(0, width, args.scale):
        dy = 1
        for y in xrange(0, height, args.scale):
            (red, green, blue) = pix[x, y]
            colorstr = "%02x%02x%02x" % (clamp(red), clamp(green), clamp(blue))
            fill = PatternFill("solid", fgColor=colorstr)
            worksheet.cell(column=dx, row=dy).fill = fill
            dy += 1
        dx += 1

    for column_cells in worksheet.columns:
        worksheet.column_dimensions[column_cells[0].column].width = COLUMN_WIDTH
    for row_cells in worksheet.rows:
        worksheet.row_dimensions[row_cells[0].row].height = ROW_HEIGHT

    wb.save(args.destination)


if __name__ == "__main__":
    main()
